"""
    Json으로 변환된 api 명세를 excel로 정리
"""

import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


def generate_excel(json_path : str, output_path : str):
    with open(json_path, "r", encoding="utf-8") as f:
        api_spec = json.load(f)

    wb = Workbook()
    ws = wb.active
    ws.title = "API 명세서"

    # 헤더 정의 
    headers = ["Controller(class)", "URL", "HTTP Method", "Controller(method)", 
               "Service(class)", "Service(method)", "Dao(class)", "Dao(Method)",
               "Params", "SQL Mapper ID", "SQL Type", "Query"]
    ws.append(headers)

    # 셀 스타일(헤더)
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    # 데이터 채우기
    for api in api_spec:
        row = [
            api.get("controller_class", ""),
            api.get("url", ""),
            api.get("http_method", ""),
            api.get("controller_method", ""),
            api.get("service_class", ""),
            api.get("service_method", ""),
            api.get("dao_class", ""),
            api.get("dao_method", ""),
            api.get("params", ""),
            api.get("dao->sql_mapper_id", ""),
            api.get("sql_type", ""),
            api.get("Query", "").replace("\n", " ")[:500]
        ]
        ws.append(row)

    # 자동 너비 조정 
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(20, min(max_length + 5, 100))

    # 저장
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f'### Excel 명세서 저장 완료 -> {output_path}')


# 테스트
if __name__ == "__main__":
    generate_excel(r"output\api_spec.json", r"output\API명세서.xlsx")